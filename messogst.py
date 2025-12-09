import streamlit as st
import pandas as pd
import io
import zipfile
import requests
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ====================================================================
#  GITHUB TEMPLATE URL (UPDATED)
# ====================================================================
GITHUB_TEMPLATE_URL = (
    "https://raw.githubusercontent.com/Biswa-hack/Messo_GST/main/MESSO%20GST%20Template.xlsx"
)

# ====================================================================
#  GLOBAL COLUMN MAPPING (explicit order used later)
# ====================================================================
COLUMN_MAPPING = {
    'order_date': 'order_date',
    'sub_order_num': 'order_num',
    'hsn_code': 'hsn_code',
    'gst_rate': 'gst_rate',
    'total_taxable_sale_value': 'tcs_taxable_amount',
    'end_customer_state_new': 'end_customer_state_new',
    'quantity': 'QTY',
}

# The exact order we will write into Excel columns B:I
WRITE_COL_ORDER = [
    'order_date',
    'order_num',
    'hsn_code',
    'gst_rate',
    'tcs_taxable_amount',
    'end_customer_state_new',
    'QTY',
    'TYPE'
]

# ====================================================================
#  STATE → GST CODE MAPPING (Column J)  -- keep keys as user-friendly names
# ====================================================================
STATE_MAPPING = {
    "Jammu and Kashmir": "01-Jammu & Kashmir",
    "Jammu & Kashmir": "01-Jammu & Kashmir",
    "Himachal Pradesh": "02-Himachal Pradesh",
    "Punjab": "03-Punjab",
    "Chandigarh": "04-Chandigarh",
    "Uttarakhand": "05-Uttarakhand",
    "Haryana": "06-Haryana",
    "Delhi": "07-Delhi",
    "Rajasthan": "08-Rajasthan",
    "Uttar Pradesh": "09-Uttar Pradesh",
    "Bihar": "10-Bihar",
    "Sikkim": "11-Sikkim",
    "Arunachal Pradesh": "12-Arunachal Pradesh",
    "Nagaland": "13-Nagaland",
    "Manipur": "14-Manipur",
    "Mizoram": "15-Mizoram",
    "Tripura": "16-Tripura",
    "Meghalaya": "17-Meghalaya",
    "Megalaya": "17-Meghalaya",  # common misspelling
    "Assam": "18-Assam",
    "West Bengal": "19-West Bengal",
    "Jharkhand": "20-Jharkhand",
    "Odisha": "21-Odisha",
    "Chhattisgarh": "22-Chhattisgarh",
    "Madhya Pradesh": "23-Madhya Pradesh",
    "Gujarat": "24-Gujarat",
    "Daman and Diu": "25-Daman & Diu",
    "Daman & Diu": "25-Daman & Diu",
    "THE DADRA AND NAGAR HAVELI AND DAMAN AND DIU": "26-Dadra & Nagar Haveli & Daman & Diu",
    "Dadra & Nagar Haveli & Daman & Diu": "26-Dadra & Nagar Haveli & Daman & Diu",
    "Dadra and Nagar Haveli": "26-Dadra & Nagar Haveli & Daman & Diu",
    "Maharashtra": "27-Maharashtra",
    "Karnataka": "29-Karnataka",
    "Goa": "30-Goa",
    "Lakshadweep": "31-Lakshdweep",
    "Kerala": "32-Kerala",
    "Tamil Nadu": "33-Tamil Nadu",
    "PONDICHERRY": "34-Puducherry",
    "Puducherry": "34-Puducherry",
    "Pondicherry": "34-Puducherry",
    "Andaman and Nico.In.": "35-Andaman & Nicobar Islands",
    "ANDAMAN AND NICOBAR ISLANDS": "35-Andaman & Nicobar Islands",
    "Andaman & Nicobar Islands": "35-Andaman & Nicobar Islands",
    "Telangana": "36-Telangana",
    "Andhra Pradesh": "37-Andhra Pradesh",
    "Ladakh": "38-Ladakh",
    "Other Territory": "97-Other Territory"
}

# Normalize STATE_MAPPING keys to a dictionary keyed by normalized string
def _normalized_mapping(mapping):
    norm = {}
    for k, v in mapping.items():
        key = str(k).strip().lower().replace("\u00A0", " ")
        key = " ".join(key.split())  # collapse multiple spaces
        norm[key] = v
    return norm

NORM_STATE_MAPPING = _normalized_mapping(STATE_MAPPING)

# ====================================================================
#  LOAD TEMPLATE FROM GITHUB URL
# ====================================================================
def load_template_from_github():
    r = requests.get(GITHUB_TEMPLATE_URL)
    if r.status_code != 200:
        st.error("❌ Could not download template from GitHub. HTTP " + str(r.status_code))
        return None
    return io.BytesIO(r.content)


# ====================================================================
#  PROCESS SINGLE EXCEL FILE (Sales or Returns)
# ====================================================================
def process_file(file_data, data_type):
    df = pd.read_excel(file_data)

    # Rename columns — only keeps those present
    df_processed = df.rename(columns=COLUMN_MAPPING)

    # Ensure all expected columns exist (create with NaN if missing)
    for col in COLUMN_MAPPING.values():
        if col not in df_processed.columns:
            df_processed[col] = pd.NA

    # Build dataframe with explicit order (pre-TYPE)
    result_cols = [c for c in WRITE_COL_ORDER if c != 'TYPE']  # first 7 cols
    df_final = df_processed[result_cols].copy()

    # Convert numeric columns safely
    if 'tcs_taxable_amount' in df_final.columns:
        df_final['tcs_taxable_amount'] = pd.to_numeric(df_final['tcs_taxable_amount'], errors='coerce').fillna(0)

    if 'QTY' in df_final.columns:
        df_final['QTY'] = pd.to_numeric(df_final['QTY'], errors='coerce').fillna(0)

    # Add TYPE as a separate column (guaranteed column position at the end)
    df_final['TYPE'] = data_type

    # For returns flip sign of numeric fields
    if data_type.lower().startswith("return"):
        df_final['tcs_taxable_amount'] = df_final['tcs_taxable_amount'].abs() * -1
        df_final['QTY'] = df_final['QTY'].abs() * -1
    else:
        df_final['tcs_taxable_amount'] = df_final['tcs_taxable_amount'].abs()
        df_final['QTY'] = df_final['QTY'].abs()

    return df_final


# ====================================================================
#  PROCESS ZIP + MERGE + WRITE TO TEMPLATE
# ====================================================================
def process_zip_and_combine_data(zip_file):
    sales_data = None
    return_data = None

    # unzip
    with zipfile.ZipFile(io.BytesIO(zip_file.read())) as z:
        for name in z.namelist():
            if name.endswith((".xlsx", ".xls")):
                if "return" in name.lower() or "rtn" in name.lower():
                    return_data = z.open(name)
                else:
                    sales_data = z.open(name)

    if sales_data is None or return_data is None:
        st.error("❌ Sales and Returns files not found in ZIP. Make sure both files are present and named clearly (contain 'return' for returns).")
        return None

    df_sales = process_file(sales_data, "Sale")
    df_returns = process_file(return_data, "Return")

    df_merged = pd.concat([df_sales, df_returns], ignore_index=True)

    # Normalize incoming state values for mapping
    def _normalize_state(s):
        s = "" if pd.isna(s) else str(s)
        s = s.strip().lower().replace("\u00A0", " ")
        s = " ".join(s.split())
        return s

    df_merged['end_customer_state_norm'] = df_merged['end_customer_state_new'].apply(_normalize_state)

    # Map using normalized keys
    df_merged['J_mapped'] = df_merged['end_customer_state_norm'].map(NORM_STATE_MAPPING).fillna("")

    # Collect unmapped state variants for diagnostics
    unmapped = (
        df_merged.loc[
            (df_merged['end_customer_state_norm'] != "") & (df_merged['J_mapped'] == ""),
            'end_customer_state_new'
        ]
        .astype(str)
        .unique()
        .tolist()
    )

    # Load template from GitHub
    template_stream = load_template_from_github()
    if template_stream is None:
        return None

    wb = load_workbook(template_stream)
    if 'raw' not in wb.sheetnames:
        st.error("❌ Template workbook does not contain a sheet named 'raw'. Please check template.")
        return None
    ws = wb["raw"]

    # Clear previous B:I rows (rows starting from 3)
    for row in range(3, ws.max_row + 1):
        for col in range(2, 10):  # columns B (2) through I (9)
            ws.cell(row=row, column=col).value = None

    start_row = 3

    # Ensure df_merged has the exact WRITE_COL_ORDER columns (create missing if needed)
    for col in WRITE_COL_ORDER:
        if col not in df_merged.columns:
            df_merged[col] = "" if col != 'QTY' and col != 'tcs_taxable_amount' else 0

    # Write B:I (explicit order)
    write_df = df_merged[WRITE_COL_ORDER]  # columns in the exact order we want
    for r_idx, row in enumerate(dataframe_to_rows(write_df, index=False, header=False)):
        for c_idx, value in enumerate(row):
            ws.cell(start_row + r_idx, 2 + c_idx).value = value

    # Column A = "Messo"
    for r in range(len(df_merged)):
        ws.cell(start_row + r, 1).value = "Messo"

    # Insert Column J (mapped) at column 10 (J)
    for r in range(len(df_merged)):
        ws.cell(start_row + r, 10).value = df_merged.loc[r, "J_mapped"]

    # Insert formulas for K–P (columns 11..16)
    # Provide six formula strings (one per column K,L,M,N,O,P). Use None when you want an empty cell.
    RAW_FORMULAS_K_TO_P = [
        None,  # K
        '=IF(J{0}=$X$22,F{0}*E{0}/100/2,0)',  # L
        '=IF(J{0}=$X$22,F{0}*E{0}/100/2,0)',  # M
        '=IF(J{0}=$X$22,0,F{0}*E{0}/100)',     # N
        '=K{0}+L{0}+M{0}+F{0}',                # O
        '=(M{0}+L{0}+K{0})/F{0}'               # P
    ]

    for r in range(len(df_merged)):
        excel_row = start_row + r
        for col_offset, formula in enumerate(RAW_FORMULAS_K_TO_P):
            if formula is not None:
                ws.cell(excel_row, 11 + col_offset).value = formula.format(excel_row)

    # Save to BytesIO and return bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), unmapped


# ====================================================================
#  STREAMLIT UI
# ====================================================================
st.set_page_config(page_title="TCS Processor", layout="wide")

st.title("📊 TCS Data Integration & Template Filler (GitHub Template Auto-loaded)")
st.markdown("---")

zipped_files = st.file_uploader("Upload ZIP containing Sales & Returns Excel files", type=["zip"])

if zipped_files:
    if st.button("🚀 Generate Report"):
        with st.spinner("Processing..."):
            result = process_zip_and_combine_data(zipped_files)

        if result:
            file_bytes, unmapped_states = result
            st.download_button(
                "⬇ Download Modified_Combo_Report.xlsx",
                file_bytes,
                "Modified_Combo_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("Done!")

            # Show diagnostics for unmapped states
            if unmapped_states:
                st.warning("Some state names were not mapped to state codes. Add them to STATE_MAPPING if needed.")
                st.write("Unmapped state values (unique):")
                st.write(unmapped_states)
            else:
                st.info("All state values were mapped successfully.")
