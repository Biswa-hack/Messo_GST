import streamlit as st
import pandas as pd
import io
import zipfile
import requests
import json
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
#  CONFIGURATION & INITIALIZATION
# ============================================================

# Set Streamlit Page Configuration
st.set_page_config(
    page_title="GSTR-1 Report Generator ✨",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Initialize Session State for persistent results and conditional rendering
if 'combo_result' not in st.session_state:
    st.session_state.combo_result = None
    st.session_state.b2cs_result = None
    st.session_state.hsn_result = None
    st.session_state.json_result = None
    st.session_state.file_name = None
    st.session_state.dynamic_gstin = "N/A"
    st.session_state.dynamic_fp = "N/A"
    st.session_state.default_state_code_numeric = "N/A"


# ============================================================
#  GLOBAL MAPPING & CONSTANTS
# ============================================================
GITHUB_TEMPLATE_URL = "https://raw.githubusercontent.com/Biswa-hack/Messo_GST/main/MESSO%20GST%20Template.xlsx"

COLUMN_MAPPING = {
    'order_date': 'order_date',
    'sub_order_num': 'order_num',
    'hsn_code': 'hsn_code',
    'gst_rate': 'gst_rate',
    'total_taxable_sale_value': 'tcs_taxable_amount',
    'end_customer_state_new': 'end_customer_state_new',
    'quantity': 'QTY'
}

WRITE_COL_ORDER = [
    'order_date',             # B
    'order_num',              # C
    'hsn_code',               # D
    'gst_rate',               # E
    'tcs_taxable_amount',     # F
    'end_customer_state_new', # G
    'TYPE',                   # H
    'QTY'                     # I
]

STATE_MAPPING = {
    "Jammu And Kashmir": "01-Jammu & Kashmir", "Jammu & Kashmir": "01-Jammu & Kashmir",
    "Himachal Pradesh": "02-Himachal Pradesh", "Punjab": "03-Punjab",
    "Chandigarh": "04-Chandigarh", "Uttarakhand": "05-Uttarakhand",
    "Haryana": "06-Haryana", "Delhi": "07-Delhi",
    "Rajasthan": "08-Rajasthan", "Uttar Pradesh": "09-Uttar Pradesh",
    "Bihar": "10-Bihar", "Sikkim": "11-Sikkim",
    "Arunachal Pradesh": "12-Arunachal Pradesh", "Nagaland": "13-Nagaland",
    "Manipur": "14-Manipur", "Mizoram": "15-Mizoram",
    "Tripura": "16-Tripura", "Megalaya": "17-Meghalaya",
    "Meghalaya": "17-Meghalaya", "Assam": "18-Assam",
    "West Bengal": "19-West Bengal", "Jharkhand": "20-Jharkhand",
    "Odisha": "21-Odisha", "Chhattisgarh": "22-Chhattisgarh",
    "Madhya Pradesh": "23-Madhya Pradesh", "Gujarat": "24-Gujarat",
    "Daman And Diu": "25-Daman & Diu", "Daman & Diu": "25-Daman & Diu",
    "The Dadra And Nagar Haveli And Daman And Diu": "26-Dadra & Nagar Haveli & Daman & Diu",
    "Dadra & Nagar Haveli & Daman & Diu": "26-Dadra & Nagar Haveli & Daman & Diu",
    "Dadra And Nagar Haveli": "26-Dadra & Nagar Haveli & Daman & Diu",
    "Maharashtra": "27-Maharashtra", "Karnataka": "29-Karnataka",
    "Goa": "30-Goa", "Lakshadweep": "31-Lakshdweep",
    "Kerala": "32-Kerala", "Tamil Nadu": "33-Tamil Nadu",
    "Pondicherry": "34-Puducherry", "Puducherry": "34-Puducherry",
    "Andaman And Nico.In.": "35-Andaman & Nicobar Islands",
    "Andaman And Nicobar Islands": "35-Andaman & Nicobar Islands",
    "Andaman & Nicobar Islands": "35-Andaman & Nicobar Islands",
    "Telangana": "36-Telangana", "Andhra Pradesh": "37-Andhra Pradesh",
    "Ladakh": "38-Ladakh", "Other Territory": "97-Other Territory"
}


# ============================================================
#  HELPER FUNCTIONS
# ============================================================
def load_template_from_github():
    """Downloads the Excel template from the specified GitHub URL."""
    r = requests.get(GITHUB_TEMPLATE_URL)
    if r.status_code != 200:
        st.error("❌ Could not download template from GitHub. Status Code: " + str(r.status_code))
        return None
    return io.BytesIO(r.content)

def process_file(file_data, data_type):
    """Reads Excel, renames columns, and adjusts values for Sales/Return."""
    df = pd.read_excel(file_data)
    df_processed = df.rename(columns=COLUMN_MAPPING)

    # Filter and create the final DataFrame with required columns
    df_final = df_processed[list(COLUMN_MAPPING.values())].copy()
    df_final["TYPE"] = data_type

    df_final["tcs_taxable_amount"] = pd.to_numeric(df_final["tcs_taxable_amount"], errors="coerce")
    df_final["QTY"] = pd.to_numeric(df_final["QTY"], errors="coerce")

    # Apply sign convention based on data type
    if data_type == "Return":
        df_final["tcs_taxable_amount"] = df_final["tcs_taxable_amount"].abs() * -1
        df_final["QTY"] = df_final["QTY"].abs() * -1
    else:
        df_final["tcs_taxable_amount"] = df_final["tcs_taxable_amount"].abs()
        df_final["QTY"] = df_final["QTY"].abs()

    return df_final

def calculate_tax_components(df, supplier_state_code_numeric):
    """
    Calculates CGST, SGST, IGST based on the dynamically provided supplier state code.
    """
    df_taxed = df.copy()
    
    # J_mapped starts with the state code (e.g., '27-Maharashtra'). Extract the numeric code.
    df_taxed["customer_state_code_numeric"] = df_taxed["J_mapped"].str[:2]
    
    # Check if Place of Supply is the same as Supplier State Code (Intra-State)
    is_intra_state = df_taxed["customer_state_code_numeric"] == supplier_state_code_numeric
    
    df_taxed["gst_rate"] = pd.to_numeric(df_taxed["gst_rate"], errors='coerce').fillna(0)
    
    total_tax_rate = df_taxed["gst_rate"] / 100
    total_tax = df_taxed["tcs_taxable_amount"] * total_tax_rate
    
    # Use where() for conditional assignment (0 if not applicable)
    df_taxed["CGST"] = total_tax.where(is_intra_state, 0) / 2
    df_taxed["SGST"] = total_tax.where(is_intra_state, 0) / 2
    df_taxed["IGST"] = total_tax.where(~is_intra_state, 0)
    
    df_taxed["Total Tax"] = df_taxed["CGST"] + df_taxed["SGST"] + df_taxed["IGST"]
    df_taxed["Total Value"] = df_taxed["tcs_taxable_amount"] + df_taxed["Total Tax"]
    
    return df_taxed

def generate_combo_excel(df_merged, template_stream):
    """Fills the raw data into the Excel template and returns the bytes."""
    
    # template_stream is a BytesIO object, need to load workbook from it
    wb = load_workbook(template_stream)
    ws = wb["raw"]

    # Clear old data
    # Note: Clearing only max 1000 rows for safety/performance, adjust if needed
    for row in range(3, 1003): 
        for col in range(1, 16):
            if ws.cell(row=row, column=col).value is not None:
                ws.cell(row=row, column=col).value = None
            else:
                break # Stop clearing when first empty row is found

    start_row = 3
    num_rows = len(df_merged)
    
    template_output = io.BytesIO()
    
    if num_rows == 0:
        # If no data, return empty excel content (optional, but safe)
        wb.save(template_output)
        return template_output.getvalue()

    # Insert B → I
    # Use iloc for safer access after filtering/concat
    write_df = df_merged[WRITE_COL_ORDER].reset_index(drop=True) 
    
    for r_idx, row in enumerate(dataframe_to_rows(write_df, index=False, header=False)):
        if r_idx < num_rows: # Safety check
            for c_idx, value in enumerate(row):
                ws.cell(start_row + r_idx, 2 + c_idx).value = value

    # Insert Column A = Messo & Column J (Mapped State Code)
    for r in range(num_rows):
        excel_row = start_row + r
        # Access mapped state using loc on the re-indexed df
        mapped_state = write_df.loc[r, df_merged.columns.get_loc("J_mapped")] 
        
        ws.cell(excel_row, 1).value = "Messo"
        ws.cell(excel_row, 10).value = mapped_state

        # Insert formulas K–O (Assuming $X$22 in the template holds the full state code string)
        # These formulas rely on values written to F (taxable value) and E (rate)
        ws.cell(excel_row, 11).value = f'=IF(J{excel_row}=$X$22,F{excel_row}*E{excel_row}/100/2,0)'
        ws.cell(excel_row, 12).value = f'=IF(J{excel_row}=$X$22,F{excel_row}*E{excel_row}/100/2,0)'
        ws.cell(excel_row, 13).value = f'=IF(J{excel_row}<>$X$22,F{excel_row}*E{excel_row}/100,0)'
        ws.cell(excel_row, 14).value = f'=K{excel_row}+L{excel_row}+M{excel_row}+F{excel_row}'
        ws.cell(excel_row, 15).value = f'=(K{excel_row}+L{excel_row}+M{excel_row})/F{excel_row}'

    wb.save(template_output)
    return template_output.getvalue()


# ============================================================
#  SUMMARY GENERATION FUNCTIONS
# ============================================================

def generate_b2cs_csv(df_merged_taxed):
    """Generates the GSTR-1 B2CS (Table 7) summary in CSV format."""
    
    summary_df = df_merged_taxed.groupby(["J_mapped", "gst_rate"]).agg(
        Taxable_Value=('tcs_taxable_amount', 'sum')
    ).reset_index()

    summary_df['Type'] = 'OE'
    summary_df['Place Of Supply'] = summary_df['J_mapped']
    summary_df['Rate'] = summary_df['gst_rate']
    summary_df['Applicable % of Tax Rate'] = ''
    summary_df['Cess Amount'] = 0.0
    summary_df['E-Commerce GSTIN'] = ''
    
    final_b2cs_df = summary_df[[
        'Type', 'Place Of Supply', 'Rate', 'Applicable % of Tax Rate',
        'Taxable_Value', 'Cess Amount', 'E-Commerce GSTIN'
    ]].rename(columns={'Taxable_Value': 'Taxable Value'})
    
    csv_output = final_b2cs_df.to_csv(index=False).encode('utf-8')
    return csv_output

def generate_hsn_summary(df_merged_taxed):
    """Generates the GSTR-1 HSN Summary (Table 12) in CSV format."""

    summary_df = df_merged_taxed.groupby(["hsn_code", "gst_rate"]).agg(
        Total_Quantity=('QTY', 'sum'),
        Total_Value=('Total Value', 'sum'),
        Total_Taxable_Value=('tcs_taxable_amount', 'sum'),
        Integrated_Tax_Amount=('IGST', 'sum'),
        Central_Tax_Amount=('CGST', 'sum'),
        State_UT_Tax_Amount=('SGST', 'sum')
    ).reset_index()

    summary_df['Description'] = ''
    summary_df['UQC'] = 'NOS-NUMBERS'
    summary_df['Cess Amount'] = 0.0

    final_hsn_df = summary_df[[
        'hsn_code', 'Description', 'UQC', 'Total_Quantity',
        'Total_Value', 'Total_Taxable_Value', 'Integrated_Tax_Amount',
        'Central_Tax_Amount', 'State_UT_Tax_Amount', 'Cess Amount', 'gst_rate'
    ]].rename(columns={
        'hsn_code': 'HSN',
        'Total_Quantity': 'Total Quantity',
        'Total_Value': 'Total Value',
        'Total_Taxable_Value': 'Taxable Value',
        'Integrated_Tax_Amount': 'Integrated Tax Amount',
        'Central_Tax_Amount': 'Central Tax Amount',
        'State_UT_Tax_Amount': 'State/UT Tax Amount',
        'gst_rate': 'Rate'
    })
    
    # Convert to CSV
    csv_output = final_hsn_df.to_csv(index=False).encode('utf-8')
    return csv_output


def generate_gstr1_json(df_merged_taxed, dynamic_gstin, dynamic_fp, supplier_state_code_numeric):
    """
    Generates the GSTR-1 JSON file structure (Table 7 B2CS and Table 12 HSN)
    using the strict schema required by the GST portal (based on user feedback).
    """
    
    # --- 1. B2CS JSON Structure (Table 7) - FLATTENED ---
    # Group by POS and Rate
    b2cs_grouped = df_merged_taxed.groupby(['J_mapped', 'gst_rate'])

    b2cs_json_list = []
    for (pos_name, rate), group in b2cs_grouped:
        
        pos_code_only = pos_name[:2] # State Code from 'XX-State Name'
        
        group_txval = group['tcs_taxable_amount'].sum()
        group_iamt = group['IGST'].sum()
        group_camt = group['CGST'].sum()
        group_samt = group['SGST'].sum()

        # Skip if Taxable Value is very close to zero
        if abs(group_txval) < 0.005: continue
            
        # Determine Supply Type (sply_ty): INTRA if POS is same as Supplier State Code, else INTER
        if pos_code_only == supplier_state_code_numeric:
            supply_type = "INTRA"
        else:
            supply_type = "INTER"
        
        # Build the B2CS transaction object (FLAT STRUCTURE REQUIRED BY PORTAL)
        b2cs_entry = {
            "sply_ty": supply_type,
            "rt": int(rate),
            "typ": "OE", # Other than E-Commerce
            "pos": pos_code_only,
            "txval": round(group_txval, 2),
            "iamt": round(group_iamt, 2),
            "camt": round(group_camt, 2),
            "samt": round(group_samt, 2),
            "csamt": 0.0
        }
        b2cs_json_list.append(b2cs_entry)


    # --- 2. HSN Summary JSON Structure (Table 12) ---
    hsn_grouped = df_merged_taxed.groupby(['hsn_code', 'gst_rate']).agg(
        qty=('QTY', 'sum'),
        txval=('tcs_taxable_amount', 'sum'),
        iamt=('IGST', 'sum'),
        camt=('CGST', 'sum'),
        samt=('SGST', 'sum')
    ).reset_index()
    
    hsn_data_list = []
    num_counter = 1
    for index, row in hsn_grouped.iterrows():
        # Ensure HSN and Rate are valid before adding
        if pd.notna(row['hsn_code']) and str(row['hsn_code']).strip() and row['gst_rate'] > 0:
            hsn_entry = {
                "num": num_counter,
                "hsn_sc": str(int(row['hsn_code'])),
                "desc": "", 
                "uqc": "NOS", # Changed from 'NOS-NUMBERS' to 'NOS' to match working sample
                "qty": round(row['qty'], 3),
                # Removed 'val' (Total Value) as per working sample
                "txval": round(row['txval'], 2),
                "iamt": round(row['iamt'], 2),
                "camt": round(row['camt'], 2),
                "samt": round(row['samt'], 2),
                "csamt": 0.0,
                "rt": int(row['gst_rate']),
            }
            hsn_data_list.append(hsn_entry)
            num_counter += 1

    # --- 3. Combine into Final GSTR-1 JSON Structure ---
    
    gstr1_json_output = {
        "gstin": dynamic_gstin,
        "fp": dynamic_fp,
        "version": "GST3.2.3", # Mandatory field added
        "hash": "hash", # Mandatory field added (placeholder)
        # Removed 'gt' and 'cur_gt' to match working sample
        "b2cs": b2cs_json_list,
        "hsn": {
            "hsn_b2c": hsn_data_list # Key changed from 'data' to 'hsn_b2c'
        }
    }
    
    # Use standard json.dumps for strict compliance
    return json.dumps(gstr1_json_output, indent=4).encode('utf-8')


# ============================================================
#  MAIN ZIP PROCESSOR
# ============================================================
def process_zip_and_combine_data(zip_file):
    """
    Extracts, processes, merges data, fills the Excel template, and generates reports.
    Sales file is mandatory for configuration; Return file is optional.
    """
    sales_data_bytes = None
    return_data_bytes = None

    # 1. Extract file streams from ZIP
    try:
        with zipfile.ZipFile(io.BytesIO(zip_file.read())) as z:
            for name in z.namelist():
                if name.endswith((".xlsx", ".xls")):
                    if "return" in name.lower() or "rtn" in name.lower():
                        return_data_bytes = z.read(name)
                    elif "sale" in name.lower() or "sls" in name.lower() or "invoice" in name.lower():
                        sales_data_bytes = z.read(name)
    except zipfile.BadZipFile:
        st.error("❌ Invalid or corrupted ZIP file.")
        return False
        
    # **Sales file is mandatory for configuration (GSTIN/FP)**
    if not sales_data_bytes:
        st.error("❌ The **Sales file** is mandatory as it contains the required configuration data (GSTIN in C2, Month/Year in P2/O2) needed for processing and file naming.")
        return False

    sales_data_stream = io.BytesIO(sales_data_bytes)

    # 1a. Extract GSTIN and Reporting Period (C2, P2, O2)
    try:
        wb_sales = load_workbook(sales_data_stream)
        ws_sales = wb_sales.active
        
        dynamic_gstin = str(ws_sales['C2'].value).strip() if ws_sales['C2'].value is not None else None
        reporting_month = ws_sales['P2'].value
        reporting_year = ws_sales['O2'].value

        sales_data_stream.seek(0) # Reset stream pointer for pandas processing below
        
        if not (dynamic_gstin and len(dynamic_gstin) == 15):
             st.error("❌ GSTIN in C2 is invalid or missing.")
             return False
        if not (reporting_month and reporting_year):
            st.error("❌ Reporting Month (P2) or Year (O2) is missing.")
            return False
        
        # Format FP and Filename
        month_str = str(reporting_month).zfill(2)
        year_str = str(reporting_year)
        if len(year_str) == 2:
            year_str = '20' + year_str
        
        dynamic_fp = f"{month_str}{year_str}"
        dynamic_filename = f"{dynamic_gstin}_{month_str}_{year_str}_GSTR1.xlsx"
        default_state_code_numeric = dynamic_gstin[:2]
        
    except Exception as e:
        st.error(f"❌ Error extracting header data from Sales file (C2, P2, O2): {e}")
        return False

    # 2. Process DataFrames
    try:
        df_sales = process_file(sales_data_stream, "Sale")
        
        # Handle optional Returns file
        if return_data_bytes:
            return_data_stream = io.BytesIO(return_data_bytes)
            df_returns = process_file(return_data_stream, "Return")
        else:
            st.warning("⚠️ Return file not found in ZIP. Processing Sales data only.")
            # Create an empty DataFrame with the expected structure for safe concatenation
            expected_cols = list(COLUMN_MAPPING.values()) + ["TYPE"]
            df_returns = pd.DataFrame(columns=expected_cols)

    except Exception as e:
        st.error(f"❌ Error processing input files: {e}")
        return False
        
    # 3. Merge DataFrames
    df_merged = pd.concat([df_sales, df_returns], ignore_index=True)

    # Map State Code
    df_merged["end_customer_state_new"] = df_merged["end_customer_state_new"].astype(str).str.title()
    df_merged["J_mapped"] = df_merged["end_customer_state_new"].map(STATE_MAPPING).fillna("")

    # 4. Calculate Tax Components 
    df_merged_taxed = calculate_tax_components(df_merged.copy(), default_state_code_numeric) 

    # 5. Load Template (for Excel output only)
    template_stream = load_template_from_github()
    if template_stream is None:
        return False

    # 6. Generate All Reports
    combo_excel_output = generate_combo_excel(df_merged, template_stream) 
    b2cs_csv_output = generate_b2cs_csv(df_merged_taxed)
    hsn_csv_output = generate_hsn_summary(df_merged_taxed)
    
    # CRITICAL: Passing the supplier state code numeric for JSON's sply_ty calculation
    json_output = generate_gstr1_json(df_merged_taxed, dynamic_gstin, dynamic_fp, default_state_code_numeric) 

    # 7. Save outputs to session state
    st.session_state.combo_result = combo_excel_output
    st.session_state.b2cs_result = b2cs_csv_output
    st.session_state.hsn_result = hsn_csv_output
    st.session_state.json_result = json_output
    st.session_state.file_name = dynamic_filename
    st.session_state.dynamic_gstin = dynamic_gstin
    st.session_state.dynamic_fp = dynamic_fp
    st.session_state.default_state_code_numeric = default_state_code_numeric
    
    return True


# ============================================================
#  STREAMLIT UI
# ============================================================

st.title("GST Report Generator ✨")
st.markdown("---")

# 🎨 Attractive UI Section
st.markdown("### ⚙️ Extracted Configuration")

# Display dynamically extracted values
st.info(f"""
**Supplier GSTIN (C2):** `{st.session_state.dynamic_gstin}`

**Reporting Period (P2/O2):** `{st.session_state.dynamic_fp}` (MMYYYY)

**Intra-State Code (from GSTIN):** `{st.session_state.default_state_code_numeric}`
""")

st.markdown("---")
st.markdown("### 📤 File Upload")

# Clear session state if a new file is uploaded
zipped_files = st.file_uploader("Upload ZIP containing Sales (Mandatory) + Return (Optional) files", type=["zip"], on_change=lambda: [
    st.session_state.update(combo_result=None, b2cs_result=None, hsn_result=None, json_result=None, file_name=None)
])

# Process button
if zipped_files:
    if st.button("🚀 Generate All 4 Reports", type="primary"):
        with st.spinner("Processing... Generating Combo, B2CS Summary (CSV), HSN Summary (CSV), and GSTR-1 JSON."):
            success = process_zip_and_combine_data(zipped_files)

        if success:
            st.success("✔️ Processing Complete! All four reports are ready for download.")

# Conditional Download Section (Visible only if session state has results)
if st.session_state.combo_result and st.session_state.b2cs_result and st.session_state.hsn_result and st.session_state.json_result:
    
    st.markdown("---")
    st.markdown("### ⬇️ Download Reports (All ready for GSTR-1 Filing)")
    st.markdown(f"**Base File Name:** `{st.session_state.file_name.replace('.xlsx', '')}`")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown("#### 1. Raw Combo Data")
        st.download_button(
            "⬇ Combo Report (.xlsx)",
            st.session_state.combo_result,
            st.session_state.file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with col2:
        st.markdown("#### 2. B2CS Summary (CSV)")
        st.download_button(
            "⬇ B2CS Summary (.csv)",
            st.session_state.b2cs_result,
            "B2CS_Summary_Report.csv",
            mime="text/csv"
        )

    with col3:
        st.markdown("#### 3. HSN Summary (CSV)")
        st.download_button(
            "⬇ HSN Summary (.csv)",
            st.session_state.hsn_result,
            "HSN_Summary_Report.csv",
            mime="text/csv"
        )

    with col4:
        st.markdown("#### 4. GSTR-1 JSON (Filing)")
        st.download_button(
            "⬇ GSTR1 JSON File",
            st.session_state.json_result,
            f"{st.session_state.file_name.replace('.xlsx', '')}_GSTR1.json",
            mime="application/json"
        )
