import streamlit as st
import pandas as pd
import io
import zipfile
import requests
import json 
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
#  CONFIGURATION & INITIALIZATION
# ============================================================

st.set_page_config(
    page_title="GSTR-1 Report Generator ✨", 
    layout="wide", 
    initial_sidebar_state="collapsed"
)

if 'combo_result' not in st.session_state:
    st.session_state.combo_result = None
    st.session_state.b2cs_result = None
    st.session_state.hsn_result = None
    st.session_state.json_result = None 
    st.session_state.file_name = None
    st.session_state.dynamic_gstin = "N/A" 
    st.session_state.dynamic_fp = "N/A"
    st.session_state.default_state_code_numeric = "N/A"

# ============================================================
#  GLOBAL MAPPING & CONSTANTS 
# ============================================================
GITHUB_TEMPLATE_URL = "https://raw.githubusercontent.com/Biswa-hack/Messo_GST/main/MESSO%20GST%20Template.xlsx"

COLUMN_MAPPING = {
    'order_date': 'order_date',
    'sub_order_num': 'order_num',
    'hsn_code': 'hsn_code',
    'gst_rate': 'gst_rate',
    'total_taxable_sale_value': 'tcs_taxable_amount',
    'end_customer_state_new': 'end_customer_state_new',
    'quantity': 'QTY'
}

WRITE_COL_ORDER = ['order_date', 'order_num', 'hsn_code', 'gst_rate', 'tcs_taxable_amount', 'end_customer_state_new', 'TYPE', 'QTY']

STATE_MAPPING = {
    "Jammu And Kashmir": "01-Jammu & Kashmir", "Jammu & Kashmir": "01-Jammu & Kashmir",
    "Himachal Pradesh": "02-Himachal Pradesh", "Punjab": "03-Punjab",
    "Chandigarh": "04-Chandigarh", "Uttarakhand": "05-Uttarakhand",
    "Haryana": "06-Haryana", "Delhi": "07-Delhi",
    "Rajasthan": "08-Rajasthan", "Uttar Pradesh": "09-Uttar Pradesh",
    "Bihar": "10-Bihar", "Sikkim": "11-Sikkim",
    "Arunachal Pradesh": "12-Arunachal Pradesh", "Nagaland": "13-Nagaland",
    "Manipur": "14-Manipur", "Mizoram": "15-Mizoram",
    "Tripura": "16-Tripura", "Meghalaya": "17-Meghalaya",
    "Assam": "18-Assam", "West Bengal": "19-West Bengal",
    "Jharkhand": "20-Jharkhand", "Odisha": "21-Odisha",
    "Chhattisgarh": "22-Chhattisgarh", "Madhya Pradesh": "23-Madhya Pradesh",
    "Gujarat": "24-Gujarat", "Daman And Diu": "25-Daman & Diu",
    "Dadra & Nagar Haveli & Daman & Diu": "26-Dadra & Nagar Haveli & Daman & Diu",
    "Maharashtra": "27-Maharashtra", "Karnataka": "29-Karnataka",
    "Goa": "30-Goa", "Lakshadweep": "31-Lakshdweep",
    "Kerala": "32-Kerala", "Tamil Nadu": "33-Tamil Nadu",
    "Puducherry": "34-Puducherry", "Andaman & Nicobar Islands": "35-Andaman & Nicobar Islands",
    "Telangana": "36-Telangana", "Andhra Pradesh": "37-Andhra Pradesh",
    "Ladakh": "38-Ladakh"
}

# ============================================================
#  HELPER FUNCTIONS
# ============================================================
def load_template_from_github():
    try:
        r = requests.get(GITHUB_TEMPLATE_URL)
        return io.BytesIO(r.content) if r.status_code == 200 else None
    except:
        return None

def process_file(file_data, data_type):
    df = pd.read_excel(file_data)
    df_processed = df.rename(columns=COLUMN_MAPPING)
    df_final = df_processed[list(COLUMN_MAPPING.values())].copy()
    df_final["TYPE"] = data_type
    df_final["tcs_taxable_amount"] = pd.to_numeric(df_final["tcs_taxable_amount"], errors="coerce").fillna(0)
    df_final["QTY"] = pd.to_numeric(df_final["QTY"], errors="coerce").fillna(0)

    if data_type == "Return":
        df_final["tcs_taxable_amount"] = df_final["tcs_taxable_amount"].abs() * -1
        df_final["QTY"] = df_final["QTY"].abs() * -1
    return df_final

def calculate_tax_components(df, supplier_state_code_numeric):
    df_taxed = df.copy() 
    df_taxed["customer_state_code_numeric"] = df_taxed["J_mapped"].str[:2]
    is_intra_state = df_taxed["customer_state_code_numeric"] == supplier_state_code_numeric
    df_taxed["gst_rate"] = pd.to_numeric(df_taxed["gst_rate"], errors='coerce').fillna(0)
    
    total_tax = df_taxed["tcs_taxable_amount"] * (df_taxed["gst_rate"] / 100)
    
    df_taxed["CGST"] = total_tax.where(is_intra_state, 0) / 2
    df_taxed["SGST"] = total_tax.where(is_intra_state, 0) / 2
    df_taxed["IGST"] = total_tax.where(~is_intra_state, 0)
    df_taxed["Total Value"] = df_taxed["tcs_taxable_amount"] + total_tax
    return df_taxed

# ============================================================
#  REVISED JSON GENERATOR (The Core Fix)
# ============================================================
def generate_gstr1_json(df_merged_taxed, dynamic_gstin, dynamic_fp):
    """
    Generates GSTR-1 JSON. 
    Fixes: Correct key names ('typ', 'det'), adds versioning, handles Intra/Inter tax splits.
    """
    # 1. B2CS (Table 7)
    b2cs_grouped = df_merged_taxed.groupby(['J_mapped', 'gst_rate'])
    b2cs_list = []

    for (pos_name, rate), group in b2cs_grouped:
        pos_code = pos_name[:2]
        txval = group['tcs_taxable_amount'].sum()
        if abs(txval) < 0.01: continue

        b2cs_list.append({
            "pos": pos_code,
            "typ": "OE",
            "det": [{
                "txval": round(txval, 2),
                "rt": int(rate),
                "iamt": round(group['IGST'].sum(), 2),
                "camt": round(group['CGST'].sum(), 2),
                "samt": round(group['SGST'].sum(), 2),
                "csamt": 0.0
            }]
        })

    # 2. HSN (Table 12)
    hsn_grouped = df_merged_taxed.groupby(['hsn_code', 'gst_rate']).agg({
        'QTY': 'sum', 'Total Value': 'sum', 'tcs_taxable_amount': 'sum',
        'IGST': 'sum', 'CGST': 'sum', 'SGST': 'sum'
    }).reset_index()
    
    hsn_data = []
    for i, row in hsn_grouped.iterrows():
        if pd.notna(row['hsn_code']):
            hsn_data.append({
                "num": i + 1,
                "hsn_sc": str(int(float(row['hsn_code']))),
                "desc": "CLOTHING",
                "uqc": "NOS",
                "qty": round(row['QTY'], 2),
                "val": round(row['Total Value'], 2),
                "txval": round(row['tcs_taxable_amount'], 2),
                "iamt": round(row['IGST'], 2),
                "camt": round(row['CGST'], 2),
                "samt": round(row['SGST'], 2),
                "csamt": 0.0,
                "rt": int(row['gst_rate'])
            })

    final_json = {
        "gstin": dynamic_gstin,
        "fp": dynamic_fp,
        "version": "GST3.2.3",
        "hash": "hash",
        "gt": round(df_merged_taxed['tcs_taxable_amount'].sum(), 2),
        "cur_gt": 0.0,
        "b2cs": b2cs_list,
        "hsn": {"data": hsn_data}
    }
    return json.dumps(final_json, indent=4).encode('utf-8')

# ============================================================
#  REMAINDER OF LOGIC (Summaries & Processing)
# ============================================================
def generate_b2cs_csv(df_merged_taxed):
    summary = df_merged_taxed.groupby(["J_mapped", "gst_rate"])['tcs_taxable_amount'].sum().reset_index()
    summary.columns = ['Place Of Supply', 'Rate', 'Taxable Value']
    summary['Type'] = 'OE'
    summary['Applicable % of Tax Rate'] = ''
    summary['Cess Amount'] = 0.0
    summary['E-Commerce GSTIN'] = ''
    return summary[['Type', 'Place Of Supply', 'Rate', 'Applicable % of Tax Rate', 'Taxable Value', 'Cess Amount', 'E-Commerce GSTIN']].to_csv(index=False).encode('utf-8')

def generate_hsn_summary(df_merged_taxed):
    hsn = df_merged_taxed.groupby(["hsn_code", "gst_rate"]).agg({
        'QTY': 'sum', 'Total Value': 'sum', 'tcs_taxable_amount': 'sum',
        'IGST': 'sum', 'CGST': 'sum', 'SGST': 'sum'
    }).reset_index()
    hsn.columns = ['HSN', 'Rate', 'Total Quantity', 'Total Value', 'Taxable Value', 'Integrated Tax Amount', 'Central Tax Amount', 'State/UT Tax Amount']
    hsn['Description'] = 'CLOTHING'
    hsn['UQC'] = 'NOS-NUMBERS'
    hsn['Cess Amount'] = 0.0
    return hsn[['HSN', 'Description', 'UQC', 'Total Quantity', 'Total Value', 'Taxable Value', 'Integrated Tax Amount', 'Central Tax Amount', 'State/UT Tax Amount', 'Cess Amount', 'Rate']].to_csv(index=False).encode('utf-8')

def generate_combo_excel(df_merged, template_stream):
    wb = load_workbook(template_stream)
    ws = wb["raw"]
    for r_idx, row in enumerate(dataframe_to_rows(df_merged[WRITE_COL_ORDER], index=False, header=False)):
        for c_idx, value in enumerate(row):
            ws.cell(3 + r_idx, 2 + c_idx).value = value
        ws.cell(3 + r_idx, 1).value = "Messo"
        ws.cell(3 + r_idx, 10).value = df_merged.loc[r_idx, "J_mapped"]
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def process_zip(zip_file):
    sales_bytes = None
    rtn_bytes = None
    with zipfile.ZipFile(io.BytesIO(zip_file.read())) as z:
        for name in z.namelist():
            if name.endswith((".xlsx", ".xls")):
                if "return" in name.lower() or "rtn" in name.lower(): rtn_bytes = z.read(name)
                else: sales_bytes = z.read(name)
    
    if not sales_bytes: return False

    wb_s = load_workbook(io.BytesIO(sales_bytes))
    ws_s = wb_s.active
    gstin = str(ws_s['C2'].value).strip()
    fp = f"{str(ws_s['P2'].value).zfill(2)}{ws_s['O2'].value}"
    
    df_s = process_file(io.BytesIO(sales_bytes), "Sale")
    df_r = process_file(io.BytesIO(rtn_bytes), "Return") if rtn_bytes else pd.DataFrame()
    df_m = pd.concat([df_s, df_r], ignore_index=True)
    df_m["J_mapped"] = df_m["end_customer_state_new"].str.title().map(STATE_MAPPING).fillna("")
    
    df_taxed = calculate_tax_components(df_m, gstin[:2])
    
    st.session_state.combo_result = generate_combo_excel(df_m, load_template_from_github())
    st.session_state.b2cs_result = generate_b2cs_csv(df_taxed)
    st.session_state.hsn_result = generate_hsn_summary(df_taxed)
    st.session_state.json_result = generate_gstr1_json(df_taxed, gstin, fp)
    st.session_state.file_name = f"{gstin}_{fp}_GSTR1.xlsx"
    st.session_state.dynamic_gstin, st.session_state.dynamic_fp = gstin, fp
    return True

# ============================================================
#  UI
# ============================================================
st.title("GSTR-1 Report Generator ✨")
st.info(f"**GSTIN:** {st.session_state.dynamic_gstin} | **Period:** {st.session_state.dynamic_fp}")

uploaded_zip = st.file_uploader("Upload ZIP (Sales + Returns)", type=["zip"])
if uploaded_zip and st.button("Generate Reports"):
    if process_zip(uploaded_zip):
        st.success("Reports Generated!")
        col1, col2, col3, col4 = st.columns(4)
        col1.download_button("Excel Combo", st.session_state.combo_result, st.session_state.file_name)
        col2.download_button("B2CS CSV", st.session_state.b2cs_result, "B2CS.csv")
        col3.download_button("HSN CSV", st.session_state.hsn_result, "HSN.csv")
        col4.download_button("GSTR-1 JSON", st.session_state.json_result, "GSTR1.json")
