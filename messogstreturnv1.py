import streamlit as st
import pandas as pd
import io
import zipfile
import requests
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
#  CONFIGURATION & INITIALIZATION
# ============================================================

# Set Streamlit Page Configuration
# Note: Streamlit themes are usually set globally, but we use wide layout and icons.
# Users can enable Dark Theme via the settings menu (top right).
st.set_page_config(
    page_title="GSTR-1 Report Generator ✨", 
    layout="wide", 
    initial_sidebar_state="collapsed"
)

# Initialize Session State for persistent results and conditional rendering
if 'combo_result' not in st.session_state:
    st.session_state.combo_result = None
    st.session_state.b2cs_result = None
    st.session_state.hsn_result = None
    st.session_state.file_name = None

# ============================================================
#  GLOBAL CONSTANTS
# ============================================================
GITHUB_TEMPLATE_URL = "https://raw.githubusercontent.com/Biswa-hack/Messo_GST/main/MESSO%20GST%20Template.xlsx"
# >>> 🎯 REQUIRED FOR DYNAMIC FILE NAMING 🎯 <<<
SUPPLIER_GSTIN = "XXAAAXXAXXX0Z0" # <<< MUST BE UPDATED BY USER'S ACTUAL GSTIN
# Default supplier state code (e.g., Maharashtra)
DEFAULT_SUPPLIER_STATE_CODE = "27-Maharashtra" 

# ============================================================
#  COLUMN MAPPING & ORDER
# ============================================================
COLUMN_MAPPING = {
    'order_date': 'order_date',
    'sub_order_num': 'order_num',
    'hsn_code': 'hsn_code',
    'gst_rate': 'gst_rate',
    'total_taxable_sale_value': 'tcs_taxable_amount',
    'end_customer_state_new': 'end_customer_state_new',
    'quantity': 'QTY'
}

WRITE_COL_ORDER = [
    'order_date',             # B
    'order_num',              # C
    'hsn_code',               # D
    'gst_rate',               # E
    'tcs_taxable_amount',     # F
    'end_customer_state_new', # G
    'TYPE',                   # H
    'QTY'                     # I
]

# ============================================================
#  STATE → GST CODE MAPPING (Column J)
# ============================================================
STATE_MAPPING = {
    "Jammu And Kashmir": "01-Jammu & Kashmir", "Jammu & Kashmir": "01-Jammu & Kashmir",
    "Himachal Pradesh": "02-Himachal Pradesh", "Punjab": "03-Punjab",
    "Chandigarh": "04-Chandigarh", "Uttarakhand": "05-Uttarakhand",
    "Haryana": "06-Haryana", "Delhi": "07-Delhi",
    "Rajasthan": "08-Rajasthan", "Uttar Pradesh": "09-Uttar Pradesh",
    "Bihar": "10-Bihar", "Sikkim": "11-Sikkim",
    "Arunachal Pradesh": "12-Arunachal Pradesh", "Nagaland": "13-Nagaland",
    "Manipur": "14-Manipur", "Mizoram": "15-Mizoram",
    "Tripura": "16-Tripura", "Megalaya": "17-Meghalaya",
    "Meghalaya": "17-Meghalaya", "Assam": "18-Assam",
    "West Bengal": "19-West Bengal", "Jharkhand": "20-Jharkhand",
    "Odisha": "21-Odisha", "Chhattisgarh": "22-Chhattisgarh",
    "Madhya Pradesh": "23-Madhya Pradesh", "Gujarat": "24-Gujarat",
    "Daman And Diu": "25-Daman & Diu", "Daman & Diu": "25-Daman & Diu",
    "The Dadra And Nagar Haveli And Daman And Diu": "26-Dadra & Nagar Haveli & Daman & Diu",
    "Dadra & Nagar Haveli & Daman & Diu": "26-Dadra & Nagar Haveli & Daman & Diu",
    "Dadra And Nagar Haveli": "26-Dadra & Nagar Haveli & Daman & Diu", 
    "Maharashtra": "27-Maharashtra", "Karnataka": "29-Karnataka",
    "Goa": "30-Goa", "Lakshadweep": "31-Lakshdweep",
    "Kerala": "32-Kerala", "Tamil Nadu": "33-Tamil Nadu",
    "Pondicherry": "34-Puducherry", "Puducherry": "34-Puducherry",
    "Andaman And Nico.In.": "35-Andaman & Nicobar Islands", 
    "Andaman And Nicobar Islands": "35-Andaman & Nicobar Islands", 
    "Andaman & Nicobar Islands": "35-Andaman & Nicobar Islands",
    "Telangana": "36-Telangana", "Andhra Pradesh": "37-Andhra Pradesh",
    "Ladakh": "38-Ladakh", "Other Territory": "97-Other Territory"
}

# ============================================================
#  HELPER FUNCTIONS 
# ============================================================
def load_template_from_github():
    """Downloads the Excel template from the specified GitHub URL."""
    r = requests.get(GITHUB_TEMPLATE_URL)
    if r.status_code != 200:
        st.error("❌ Could not download template from GitHub. Status Code: " + str(r.status_code))
        return None
    return io.BytesIO(r.content)

def process_file(file_data, data_type):
    """Reads Excel, renames columns, and adjusts values for Sales/Return."""
    df = pd.read_excel(file_data)
    df_processed = df.rename(columns=COLUMN_MAPPING)

    df_final = df_processed[list(COLUMN_MAPPING.values())].copy()
    df_final["TYPE"] = data_type

    df_final["tcs_taxable_amount"] = pd.to_numeric(df_final["tcs_taxable_amount"], errors="coerce")
    df_final["QTY"] = pd.to_numeric(df_final["QTY"], errors="coerce")

    if data_type == "Return":
        df_final["tcs_taxable_amount"] = df_final["tcs_taxable_amount"].abs() * -1
        df_final["QTY"] = df_final["QTY"].abs() * -1
    else:
        df_final["tcs_taxable_amount"] = df_final["tcs_taxable_amount"].abs()
        df_final["QTY"] = df_final["QTY"].abs()

    return df_final

def calculate_tax_components(df):
    """
    Calculates CGST, SGST, IGST, and Total Tax for the DataFrame based on 
    the Place of Supply (J_mapped) vs. the DEFAULT_SUPPLIER_STATE_CODE.
    """
    df_taxed = df.copy() 
    
    is_intra_state = df_taxed["J_mapped"] == DEFAULT_SUPPLIER_STATE_CODE
    
    df_taxed["gst_rate"] = pd.to_numeric(df_taxed["gst_rate"], errors='coerce').fillna(0)
    
    total_tax_rate = df_taxed["gst_rate"] / 100
    total_tax = df_taxed["tcs_taxable_amount"] * total_tax_rate
    
    df_taxed["CGST"] = total_tax.where(is_intra_state, 0) / 2
    df_taxed["SGST"] = total_tax.where(is_intra_state, 0) / 2
    df_taxed["IGST"] = total_tax.where(~is_intra_state, 0)
    df_taxed["Total Tax"] = df_taxed["CGST"] + df_taxed["SGST"] + df_taxed["IGST"]
    df_taxed["Total Value"] = df_taxed["tcs_taxable_amount"] + df_taxed["Total Tax"]
    
    return df_taxed

# ============================================================
#  SUMMARY GENERATION FUNCTIONS
# ============================================================

def generate_b2cs_csv(df_merged_taxed):
    """Generates the GSTR-1 B2CS (Table 7) summary in CSV format."""
    
    summary_df = df_merged_taxed.groupby(["J_mapped", "gst_rate"]).agg(
        Taxable_Value=('tcs_taxable_amount', 'sum')
    ).reset_index()

    summary_df['Type'] = 'OE' # Other than E-Commerce
    summary_df['Place Of Supply'] = summary_df['J_mapped']
    summary_df['Rate'] = summary_df['gst_rate']
    summary_df['Applicable % of Tax Rate'] = '' 
    summary_df['Cess Amount'] = ''
    summary_df['E-Commerce GSTIN'] = ''
    
    final_b2cs_df = summary_df[[
        'Type', 
        'Place Of Supply', 
        'Rate', 
        'Applicable % of Tax Rate', 
        'Taxable_Value',
        'Cess Amount', 
        'E-Commerce GSTIN'
    ]].rename(columns={'Taxable_Value': 'Taxable Value'})
    
    csv_output = final_b2cs_df.to_csv(index=False).encode('utf-8')
    return csv_output

def generate_hsn_summary(df_merged_taxed):
    """Generates the GSTR-1 HSN Summary (Table 12) in Excel format."""

    summary_df = df_merged_taxed.groupby(["hsn_code", "gst_rate"]).agg(
        Total_Quantity=('QTY', 'sum'),
        Total_Taxable_Value=('tcs_taxable_amount', 'sum'),
        Total_Value=('Total Value', 'sum'),
        Integrated_Tax_Amount=('IGST', 'sum'),
        Central_Tax_Amount=('CGST', 'sum'),
        State_UT_Tax_Amount=('SGST', 'sum')
    ).reset_index()

    summary_df['Description'] = '' 
    summary_df['UQC'] = 'NOS-NUMBERS' 
    summary_df['Cess Amount'] = 0.0

    final_hsn_df = summary_df[[
        'hsn_code', 'Description', 'UQC', 'Total_Quantity',
        'Total_Value', 'Total_Taxable_Value', 'Integrated_Tax_Amount',
        'Central_Tax_Amount', 'State_UT_Tax_Amount', 'Cess Amount', 'gst_rate'
    ]].rename(columns={
        'hsn_code': 'HSN',
        'Total_Quantity': 'Total Quantity',
        'Total_Value': 'Total Value',
        'Total_Taxable_Value': 'Taxable Value',
        'Integrated_Tax_Amount': 'Integrated Tax Amount',
        'Central_Tax_Amount': 'Central Tax Amount',
        'State_UT_Tax_Amount': 'State/UT Tax Amount',
        'gst_rate': 'Rate'
    })
    
    excel_output = io.BytesIO()
    final_hsn_df.to_excel(excel_output, index=False, sheet_name='HSN_Summary')
    return excel_output.getvalue()

# ============================================================
#  MAIN ZIP PROCESSOR
# ============================================================
def process_zip_and_combine_data(zip_file):
    """Extracts, processes, merges data, fills the Excel template, and generates summaries."""
    sales_data = None
    return_data = None

    # 1. Extract Sales and Return files from ZIP
    try:
        with zipfile.ZipFile(io.BytesIO(zip_file.read())) as z:
            for name in z.namelist():
                if name.endswith((".xlsx", ".xls")):
                    if "return" in name.lower() or "rtn" in name.lower():
                        return_data = z.open(name)
                    elif "sale" in name.lower() or "sls" in name.lower() or "invoice" in name.lower():
                        sales_data = z.open(name)
    except zipfile.BadZipFile:
        st.error("❌ Invalid or corrupted ZIP file.")
        return False
        
    if not sales_data or not return_data:
        st.error("❌ ZIP must contain both a **Sales** and a **Return** file (filenames must contain key identifying words like 'sale' and 'return').")
        return False

    # 2. Process and Merge DataFrames
    try:
        df_sales = process_file(sales_data, "Sale") 
        df_returns = process_file(return_data, "Return") 
    except Exception as e:
        st.error(f"❌ Error processing input files: {e}")
        return False
        
    df_merged = pd.concat([df_sales, df_returns], ignore_index=True)

    # Standardize State Names
    df_merged["end_customer_state_new"] = df_merged["end_customer_state_new"].str.title()
    df_merged["J_mapped"] = df_merged["end_customer_state_new"].map(STATE_MAPPING).fillna("")

    # 3. Calculate Tax Components and generate summaries
    df_merged_taxed = calculate_tax_components(df_merged.copy()) 

    b2cs_csv_output = generate_b2cs_csv(df_merged_taxed)
    hsn_excel_output = generate_hsn_summary(df_merged_taxed)

    # 4. Generate dynamic file name
    df_merged['order_date'] = pd.to_datetime(df_merged['order_date'], errors='coerce')
    max_date = df_merged['order_date'].max()
    
    if pd.notna(max_date):
        # Format: MM_YYYY
        reporting_month = max_date.strftime('%m')
        reporting_year = max_date.strftime('%Y')
        dynamic_filename = f"{SUPPLIER_GSTIN}_{reporting_month}_{reporting_year}_GSTR1.xlsx"
    else:
        dynamic_filename = "Modified_Combo_Report.xlsx"

    # 5. Load Template and Insert Raw Data (Excel Template)
    template_stream = load_template_from_github()
    if template_stream is None:
        return False

    wb = load_workbook(template_stream)
    ws = wb["raw"]

    # Clear old data
    for row in range(3, ws.max_row + 1):
        for col in range(1, 16):
            ws.cell(row=row, column=col).value = None

    start_row = 3
    num_rows = len(df_merged)

    # Insert B → I
    write_df = df_merged[WRITE_COL_ORDER]

    for r_idx, row in enumerate(dataframe_to_rows(write_df, index=False, header=False)):
        for c_idx, value in enumerate(row):
            ws.cell(start_row + r_idx, 2 + c_idx).value = value

    # Insert Column A = Messo & Column J (Mapped State Code)
    for r in range(num_rows):
        excel_row = start_row + r
        ws.cell(excel_row, 1).value = "Messo"
        ws.cell(excel_row, 10).value = df_merged.loc[r, "J_mapped"]

        # Insert formulas K–O
        # K: CGST 
        ws.cell(excel_row, 11).value = f"=IF(J{excel_row}=$X$22,F{excel_row}*E{excel_row}/100/2,0)"
        # L: SGST 
        ws.cell(excel_row, 12).value = f"=IF(J{excel_row}=$X$22,F{excel_row}*E{excel_row}/100/2,0)"
        # M: IGST 
        ws.cell(excel_row, 13).value = f"=IF(J{excel_row}<>$X$22,F{excel_row}*E{excel_row}/100,0)"
        # N: Total Value
        ws.cell(excel_row, 14).value = f"=K{excel_row}+L{excel_row}+M{excel_row}+F{excel_row}"
        # O: % GST
        ws.cell(excel_row, 15).value = f"=(K{excel_row}+L{excel_row}+M{excel_row})/F{excel_row}" 

    template_output = io.BytesIO()
    wb.save(template_output)
    
    # 6. Save outputs to session state for persistent download links
    st.session_state.combo_result = template_output.getvalue()
    st.session_state.b2cs_result = b2cs_csv_output
    st.session_state.hsn_result = hsn_excel_output
    st.session_state.file_name = dynamic_filename
    
    return True # Return success flag

# ============================================================
#  STREAMLIT UI
# ============================================================

st.title("GST Report Generator ✨")
st.markdown("---")

# 🎨 Attractive UI Section
st.markdown("### ⚙️ Configuration")
st.code(f"SUPPLIER_GSTIN = \"{SUPPLIER_GSTIN}\"")
st.info(f"**Action Required:** Please check and update the `SUPPLIER_GSTIN` constant in the code above with your actual GSTIN for correct file naming. Default State for IGST calculation is: `{DEFAULT_SUPPLIER_STATE_CODE}`.")

st.markdown("---")
st.markdown("### 📤 File Upload")

# Clear session state if a new file is uploaded
zipped_files = st.file_uploader("Upload ZIP containing Sales + Return files", type=["zip"], on_change=lambda: [
    st.session_state.update(combo_result=None, b2cs_result=None, hsn_result=None, file_name=None)
])

# Process button
if zipped_files:
    if st.button("🚀 Generate All 3 Reports", type="primary"):
        with st.spinner("Processing... Generating Combo, B2CS Summary (CSV), and HSN Summary (Excel)."):
            success = process_zip_and_combine_data(zipped_files)

        if success:
            st.success("✔️ Processing Complete! Your reports are ready for download.")
        else:
            # Error handling already done inside process_zip_and_combine_data
            pass

# Conditional Download Section (Visible only if session state has results)
if st.session_state.combo_result and st.session_state.b2cs_result and st.session_state.hsn_result:
    
    st.markdown("---")
    st.markdown("### ⬇️ Download Reports")
    st.markdown(f"**Dynamic Combo File Name:** `{st.session_state.file_name}`")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("#### 1. Raw Combo Data")
        st.download_button(
            "⬇ Combo Report (.xlsx)",
            st.session_state.combo_result,
            st.session_state.file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with col2:
        st.markdown("#### 2. B2CS Summary (GSTR-1 Table 7)")
        st.download_button(
            "⬇ B2CS Summary (.csv)",
            st.session_state.b2cs_result,
            "B2CS_Summary_Report.csv",
            mime="text/csv"
        )

    with col3:
        st.markdown("#### 3. HSN Summary (GSTR-1 Table 12)")
        st.download_button(
            "⬇ HSN Summary (.xlsx)",
            st.session_state.hsn_result,
            "HSN_Summary_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
